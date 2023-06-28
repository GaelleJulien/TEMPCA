from openpyxl import Workbook
from openpyxl import load_workbook


workbookCoords = load_workbook(filename = "donnees_moi.xlsx")

sheet = workbookCoords.active
#parcours des feuilles

for row in sheet.iter_rows():
    for cell in row:
        if cell.column != 1 and cell.offset(column=-1).value == "UserID":
            # Récupération des coordonnées de la cellule
            USERID = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {USERID}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Time in bed":
            # Récupération des coordonnées de la cellule
            TIME_IN_BED = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {TIME_IN_BED}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Assumed sleep":
            # Récupération des coordonnées de la cellule
            ASSUMED_SLEEP = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {ASSUMED_SLEEP}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Actual sleep time":
            # Récupération des coordonnées de la cellule
            ACTUAL_SLEEP_TIME = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {ACTUAL_SLEEP_TIME}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Actual wake time":
            # Récupération des coordonnées de la cellule
            ACTUAL_WAKE_TIME = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {ACTUAL_WAKE_TIME}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Actual wake (%)":
            # Récupération des coordonnées de la cellule
            ACTUAL_WAKE_RATE = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {ACTUAL_WAKE_RATE}")
        elif cell.column != 1 and cell.offset(column=-1).value== "Sleep latency":
            # Récupération des coordonnées de la cellule
            SLEEP_LATENCY = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {SLEEP_LATENCY}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Sleep bouts":
            # Récupération des coordonnées de la cellule
            SLEEP_BOUTS = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {SLEEP_BOUTS}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Wake bouts":
            # Récupération des coordonnées de la cellule
            WAKE_BOUTS = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {WAKE_BOUTS}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Mean sleep bout":
            # Récupération des coordonnées de la cellule
            MEAN_SLEEP_BOUTS = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {MEAN_SLEEP_BOUTS}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Mean wake bout":
            # Récupération des coordonnées de la cellule
            MEAN_WAKE_BOUTS = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {MEAN_WAKE_BOUTS}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Immobile bouts":
            # Récupération des coordonnées de la cellule
            IMMOBILE_BOUTS = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {IMMOBILE_BOUTS}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Mean immobile bout":
            # Récupération des coordonnées de la cellule
            MEAN_IMMOBILE_BOUTS = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {MEAN_IMMOBILE_BOUTS}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Lights out":
            # Récupération des coordonnées de la cellule
            LIGHTS_OUT = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {LIGHTS_OUT}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Fell asleep":
            # Récupération des coordonnées de la cellule
            FELL_ASLEEP = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {FELL_ASLEEP}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Woke up":
            # Récupération des coordonnées de la cellule
            WOKE_UP = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {WOKE_UP}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Got up":
            # Récupération des coordonnées de la cellule
            GOT_UP = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {GOT_UP}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Actual sleep (%)":
            # Récupération des coordonnées de la cellule
            ACTUAL_SLEEP_RATE = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {ACTUAL_SLEEP_RATE}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Sleep efficiency (%)":
            # Récupération des coordonnées de la cellule
            SLEEP_EFFICIENCY = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {SLEEP_EFFICIENCY}")
        elif cell.column != 1 and cell.offset(column=-1).value == "Fragmentation Index":
            # Récupération des coordonnées de la cellule
            FRAGMENTATION_INDEX = cell.coordinate
            print(f"La cellule '{cell.value}' a les coordonnées : {FRAGMENTATION_INDEX}")
        
        elif cell.value == "Unrounded":
            # Récupération des coordonnées de la cellule
            rowsToSkip = cell.row

#mapping




