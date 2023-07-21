
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill
from openpyxl.chart import Reference, LineChart
from openpyxl.styles import Font, Color, Alignment, Border
from openpyxl.utils.dataframe import dataframe_to_rows


from classes import Stats
from classes import ToplevelWindow, MainWindow
#from mapping import rowsToSkip, USERID, ACTUAL_SLEEP_RATE, ACTUAL_WAKE_RATE, FELL_ASLEEP, WOKE_UP, SLEEP_BOUTS, WAKE_BOUTS, IMMOBILE_BOUTS, MEAN_IMMOBILE_BOUTS, MEAN_SLEEP_BOUTS, MEAN_WAKE_BOUTS, TIME_IN_BED, SLEEP_LATENCY, ASSUMED_SLEEP, SLEEP_EFFICIENCY, ACTUAL_SLEEP_TIME, ACTUAL_WAKE_TIME, LIGHTS_OUT, FRAGMENTATION_INDEX, GOT_UP

import tkinter as tk
from tkinter import *
from tkinter import filedialog, ttk

import csv
import pandas as pd
import pandastable as pt
from pandastable import Table, TableModel, config

from stats import moy, moydf, tri, triSE, plotHypnnogramme, plotHourbyUser, moydfHour, df_sorted

from PIL import ImageTk, Image
import matplotlib.pyplot as plt



import customtkinter
customtkinter.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("green")  # Themes: "blue" (standard), "green", "dark-blue"


file_path = ""

activityDonnees = []


main_window = MainWindow()




def ajouter_point(heure, valeur):
    activityDonnees.append([heure, valeur])
        

def loadWorkbook(file_path):

    workbookCoords = load_workbook(filename = file_path)

    sheet = workbookCoords.active
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.column != 1 and cell.offset(column=-1).value == "UserID":
                USERID = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Time in bed":
                TIME_IN_BED = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Assumed sleep":
                ASSUMED_SLEEP = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Actual sleep time":
                ACTUAL_SLEEP_TIME = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Actual wake time":
                ACTUAL_WAKE_TIME = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Actual wake (%)":
                ACTUAL_WAKE_RATE = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value== "Sleep latency":
                SLEEP_LATENCY = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Sleep bouts":
                SLEEP_BOUTS = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Wake bouts":
                WAKE_BOUTS = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Mean sleep bout":
                MEAN_SLEEP_BOUTS = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Mean wake bout":
                MEAN_WAKE_BOUTS = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Immobile bouts":
                IMMOBILE_BOUTS = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Mean immobile bout":
                MEAN_IMMOBILE_BOUTS = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Lights out":
                LIGHTS_OUT = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Fell asleep":
                FELL_ASLEEP = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Woke up":
                WOKE_UP = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Got up":
                GOT_UP = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Actual sleep (%)":
                ACTUAL_SLEEP_RATE = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Sleep efficiency (%)":
                SLEEP_EFFICIENCY = cell.coordinate
            elif cell.column != 1 and cell.offset(column=-1).value == "Fragmentation Index":
                FRAGMENTATION_INDEX = cell.coordinate

            elif cell.value == "Unrounded":
                rowsToSkip = cell.row

    workbook2 = load_workbook(filename = file_path)
    workbook2.sheetnames
    stats = []
    selected_users=[]
    selected_temps=[]



    #parcours des feuilles
    for sheet in workbook2 : 

        activityDonnees = []

        df2 = pd.read_excel(file_path, sheet.title, skiprows= rowsToSkip)
        print(df2)
        heures = df2['Time']
        valeurs = df2['Activity']
        for i in range(len(heures)):
            heure = heures[i]
            valeur = valeurs[i]
            activityDonnees.append([heure, valeur])

        
        #les titres des feuilles sont au format USERID_TEMP (c'est moi qui l'ai décidé nah)
        #  --> n'ayant pas de vraibale température sur MotionWare on récupère la température de la nuit à partir du nom de la feuille
        nuit = str(sheet.title[-2:])
            # Parcours des cellules

        
            #remplissage de la classe
        stat = Stats(id = sheet[USERID].value, NUIT = nuit, SPT=sheet[ASSUMED_SLEEP].value, TST= sheet[ACTUAL_SLEEP_TIME].value, 
                 actual_sleep_rate = sheet[ACTUAL_SLEEP_RATE].value, actual_wake_time=sheet[ACTUAL_WAKE_TIME].value, actual_wake_rate = sheet[ACTUAL_WAKE_RATE].value, 
                 TIB=sheet[TIME_IN_BED].value, sleep_efficiency=sheet[SLEEP_EFFICIENCY].value,lights_out=sheet[LIGHTS_OUT].value, fell_asleep=sheet[FELL_ASLEEP].value, 
                 sleep_latency=sheet[SLEEP_LATENCY].value ,woke_up=sheet[WOKE_UP].value, got_up=sheet[GOT_UP].value, SFI=sheet[FRAGMENTATION_INDEX].value, activity=activityDonnees,
                 sleep_bouts=sheet[SLEEP_BOUTS].value, wake_bouts=sheet[WAKE_BOUTS].value, immobile_bouts=sheet[IMMOBILE_BOUTS].value, mean_immobile_bouts=sheet[MEAN_IMMOBILE_BOUTS].value,
                 mean_sleep_bouts=sheet[MEAN_SLEEP_BOUTS].value, mean_wake_bouts=sheet[MEAN_WAKE_BOUTS].value)
    
        stats.append(stat)
        heures = []
        valeurs = []


    sheet = workbook2.active
    



    workbook = Workbook()
    sheet = workbook.active

   

    #nom des en-têtes
    sheet.append(["UserID", "NUIT", "TIB", "SPT", "TST", "actual_sleep (%)", "actual_wake_time", "actual_wake (%)", "sleep_efficiency (%)", "lights_out", 
              "fell_asleep", "sleep_latency", "woke_up", "got_up", "SFI"])

    #remplissage du tableau
    for stat in stats : 
        data = [stat.id, stat.NUIT, stat.TIB, stat.SPT, stat.TST, stat.actual_sleep_rate, stat.actual_wake_time, stat.actual_wake_rate,stat.sleep_efficiency,
            stat.lights_out, stat.fell_asleep, stat.sleep_latency, stat.woke_up, stat.got_up, stat.SFI]
        sheet.append(data)

    #Ajustement de la largeur de la colonne en fonction de son contenu
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))  
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value

    #remplissage de ligne alternatif
    gris = "e2e2e2"
    for rows in sheet.iter_rows(min_row=2, max_row=69, min_col=1, max_col=15):
        for cell in rows:
            if cell.row % 2:
                cell.fill = PatternFill(start_color=gris, end_color=gris,fill_type = "solid")

    #code couleur pour la température
    color_scale_rule = ColorScaleRule(end_type="num",
                                  end_value=32,
                                  end_color="e67d59",  # Rouge
                                  mid_type="num",
                                  mid_value=24,
                                  mid_color="9edd87",  # vert
                                  start_type="num",
                                  start_value=16,
                                  start_color="c3d2ff")  # bleu
    sheet.conditional_formatting.add("B2:B69", color_scale_rule)


    #filtres
    sheet.auto_filter.ref = sheet.dimensions

    #epingle la première ligne
    sheet.freeze_panes = "B2"


    workbook.save(filename= "test.xlsx")

    workbookActivity = Workbook()
    sheetActivity = workbookActivity.active
    sheetActivity.append(["UserID", "NUIT", "sleep_bouts", "wake_bouts", "immobile_bouts", "mean_immobile_bouts", "mean_sleep_bouts", "mean_wake_bouts"])
    for stat in stats : 
        dataActivity = [stat.id, stat.NUIT, stat.sleep_bouts, stat.wake_bouts, stat.immobile_bouts, stat.mean_immobile_bouts, stat.mean_sleep_bouts, stat.mean_wake_bouts]
        sheetActivity.append(dataActivity)
    workbookActivity.save(filename = "activityData.xlsx")
    workbookActivity.close()

    df = pd.DataFrame(pd.read_excel("test.xlsx"))

    dfMoy = pd.DataFrame(pd.read_excel("test.xlsx"))
    dfMoy = dfMoy.astype({"NUIT" : str})
    dfMoy["sleep_latency" ]= dfMoy["sleep_latency"].astype("string")
    dfMoy["TST" ]= dfMoy["TST"].astype("string")
    dfMoy["TIB" ]= dfMoy["TIB"].astype("string")
    dfMoy["SPT" ]= dfMoy["SPT"].astype("string")
    dfMoy["actual_wake_time" ]= dfMoy["actual_wake_time"].astype("string")



    dfMoy["sleep_latency" ] = pd.to_timedelta(dfMoy["sleep_latency" ])
    dfMoy["TST" ] = pd.to_timedelta(dfMoy["TST" ])
    dfMoy["TIB" ] = pd.to_timedelta(dfMoy["TIB" ])
    dfMoy["SPT" ] = pd.to_timedelta(dfMoy["SPT" ])
    dfMoy["actual_wake_time" ] = pd.to_timedelta(dfMoy["actual_wake_time" ])

    df_hour = pd.DataFrame(pd.read_excel("activityData.xlsx"))

    df_hour["mean_sleep_bouts" ]= df_hour["mean_sleep_bouts"].astype("string")
    df_hour["mean_wake_bouts" ]= df_hour["mean_wake_bouts"].astype("string")

    df_hour["mean_sleep_bouts" ] = pd.to_timedelta(df_hour["mean_sleep_bouts" ])
    df_hour["mean_wake_bouts" ] = pd.to_timedelta(df_hour["mean_wake_bouts" ])

    df_hour["mean_immobile_bouts"] = df_hour["mean_immobile_bouts"].astype("string")
    df_hour["mean_immobile_bouts"] = pd.to_timedelta(df_hour["mean_immobile_bouts"])

    moydf = dfMoy

    moydf = moydf.groupby("NUIT")[["sleep_efficiency (%)", "TST","actual_sleep (%)", "actual_wake (%)", "sleep_latency", "SFI"]].mean()
    moydfHour = df_hour.groupby("NUIT")[["sleep_bouts", "wake_bouts", "mean_immobile_bouts"]].mean()
    moydfHour.reset_index(inplace=True)
    moydfHour["mean_immobile_bouts"] = moydfHour["mean_immobile_bouts"].astype(str).map(lambda x: x[7:15])


    fig, ax = plt.subplots()

    moydf.reset_index(inplace = True)
    moydf = moydf.rename(columns={"index" : "NUIT"})
    moydf["sleep_latency"] = moydf["sleep_latency"].astype(str).map(lambda x: x[7:15])
    moydf["TST"] = moydf["TST"].astype(str).map(lambda x: x[7:15])


    plotMoy = moydf.plot.bar(x = "NUIT", xlabel = "Nuit", ylabel = "????")
    plt.savefig("plotMoyennes.png")


    moy = str(moydf)
    moydf["sleep_latency"] =  pd.to_timedelta(moydf["sleep_latency" ])

    df2 = pd.read_excel("CombinedActivity.xlsx")

    
    df2['Time'] = pd.to_datetime(df2['Time'],  format="%H:%M:%S")
    
    def taux_sup_20(x):
        return (x > 20).mean()

    df2.set_index("Time", inplace=True)
    df3 = df2.groupby(pd.Grouper(freq="30Min")).apply(lambda x: (x > 20).mean() * 100)

#df3 = df2.resample('15Min', on="Time").mean()
    df3.index = pd.to_datetime(df3.index)

    df3['DateTime'] = df3.index

    # Définir la plage horaire souhaitée
    start_time = pd.to_datetime('22:00:00').time()
    end_time = pd.to_datetime('07:00:00').time()

    def custom_sort(time):
        if time.time() >= start_time:
            return time
        else:
            return time + pd.DateOffset(days=1)
    
    df3["Heure"] = df3['DateTime'].map(custom_sort)
    df_sorted = df3.sort_values(by="Heure")
    df_sorted = df_sorted.dropna()

    df_sorted.reset_index(drop=True)

    df_sorted.set_index("Heure")

    df_sorted['Heure'] = pd.to_datetime(df_sorted['Heure'], format='%H:%M:%S')

    df_sorted["Hour"] = (df_sorted["Heure"].dt.hour)

    newcolumns = ["Heure"] + [col for col in df_sorted.columns if col !="Heure" ]
    df_sorted = df_sorted[newcolumns]

    def taux_sup_20(column):
        count_sup_20 = sum(column > 20)  
        total_count = len(column)  
        taux = count_sup_20 / total_count  
        return taux


    df_hourly = df_sorted.groupby("Hour", sort=False).mean()
    df_hourly = df_hourly.drop(columns=["DateTime", "Heure"])

    
# Calculer le taux d'éveil pour chaque heure
    df_hourly['AwakeningRate'] = (df_hourly.iloc[:, 1:] >= 20).mean(axis=1) * 100  # Assuming columns from 1 represent users
    print(df_hourly)

    fig, ax = plt.subplots()
    ax.bar(df_hourly.index, df_hourly['AwakeningRate'])

    ax.set_xlabel('Heure')
    ax.set_ylabel("Taux d'éveil (%)")
    ax.set_title("Taux d'éveil par heure")
    plt.savefig("Wakefulness_Rate_per_Hour.png")


    dfBoxSE = pd.DataFrame(dfMoy[["sleep_efficiency (%)", "NUIT"]])
    boxfig = dfBoxSE.plot.box(by="NUIT", xlabel = "Nuit", ylabel = "Sleep efficiency (%)")
    plt.savefig("boxplotSE.png")


    dfBoxTimeSFI = pd.DataFrame(dfMoy[["SFI", "NUIT"]])
    boxfig = dfBoxTimeSFI.plot.box(by="NUIT", xlabel = "Nuit", ylabel = "Sleep Fragmentation Index")
    plt.savefig("boxplotSFI.png")


    dfMoy["sleep_latency" ] = dfMoy["sleep_latency" ] / pd.Timedelta(minutes=1)
    dfBoxTimeSL = pd.DataFrame(dfMoy[["sleep_latency", "NUIT"]])
    boxfig = dfBoxTimeSL.plot.box(by="NUIT", xlabel = "Nuit", ylabel = "Sleep Latency (minutes)")
    plt.savefig("boxplotSL.png")

    dfMoy = pd.DataFrame(moydf[["sleep_efficiency (%)", "NUIT"]])
    fig2 = dfMoy.plot.bar(x = "NUIT", y = "sleep_efficiency (%)", rot = 0, xlabel = "Nuit", ylabel = "Sleep efficiency (%)")
    plt.savefig("sleep_efficiency_all_means.png")

    dfMoy = pd.DataFrame(moydf[["SFI", "NUIT"]])
    fig2 = dfMoy.plot.bar(x = "NUIT", y = "SFI", rot = 0, xlabel = "Nuit", ylabel = "Sleep Fragmentation Index")
    plt.savefig("sfi_all_means.png")

    moydf["sleep_latency" ] = moydf["sleep_latency" ] / pd.Timedelta(minutes=1)
    dfMoy = pd.DataFrame(moydf[["sleep_latency", "NUIT"]])
    fig3 = dfMoy.plot.bar(x = "NUIT", y = "sleep_latency", rot = 0, xlabel = "Nuit", ylabel = "Sleep Latency (minutes)")
    plt.savefig("sleep_latency_all_means.png")

    dfMoy = pd.DataFrame(moydf)
    fig3 = dfMoy.plot.bar(x = "NUIT", rot = 0)
    plt.savefig("all_means.png")


    filter = StringVar()
    
    filter.set(moy)

    users = ["Sujets..."]
    df = df.astype({"NUIT" : str})
    users.extend(df["UserID"].drop_duplicates().to_list())
    


    temp = ["Température..."]
    temp.extend(df["NUIT"].drop_duplicates().to_list())



    def optionmenu_triSFI_callback(choice):
        if(choice == tri[0]):
            newdf = df
        if(choice == tri[1]): 
            newdf = df.sort_values(by = "SFI")
        if(choice == tri[2]):
            newdf = df.sort_values(by = "SFI", ascending=False)

        fillTable(newdf, "Données")
        
    
    def optionmenu_triSE_callback(choice):

        if(choice == triSE[0]):
            newdf = df
        if(choice == triSE[1]): 
            newdf = df.sort_values(by = "sleep_efficiency (%)")
        if(choice == triSE[2]): 
            newdf = df.sort_values(by = "sleep_efficiency (%)", ascending=False)
        fillTable(newdf, "Données")



    def checkbox_user_callback():
        if(checkbox_var_users.get() in selected_users):
            selected_users.remove(checkbox_var_users.get())
            print(selected_users)
            if(selected_temps != []):
                if(selected_users == []) :
                    filtrage = (df["NUIT"].astype(str).isin(selected_temps)) 
                    newdf_users = df[filtrage]
                else :
                    filtrage= (df["UserID"].isin(selected_users) & df["NUIT"].astype(str).isin(selected_temps))
                    newdf_users = df[filtrage]

            else : 
                if(selected_users == []) :
                    newdf_users = df
                else : 
                    filtrage= (df["UserID"].isin(selected_users))
                    newdf_users = df[filtrage]
            
        else : 
            selected_users.append(checkbox_var_users.get())
            print(selected_users)
            if(selected_temps != []):
                filtrage= (df["UserID"].isin(selected_users) & df["NUIT"].astype(str).isin(selected_temps))
            else : 
                filtrage= (df["UserID"].isin(selected_users))

            newdf_users = df[filtrage]
        
        fillTable(newdf_users, "Données")



    def checkbox_temp_callback():
        print(checkbox_var_temp.get())
        if(checkbox_var_temp.get() in selected_temps):
            selected_temps.remove(checkbox_var_temp.get())
            if(selected_users != []):
                if(selected_temps == []) :
                    filtrage = (df["UserID"].isin(selected_users)) 
                    newdf_temp = df[filtrage]
                else:
                    filtrage= (df["UserID"].isin(selected_users) & df["NUIT"].astype(str).isin(selected_temps))
                    newdf_temp = df[filtrage]


            else : 
                if(selected_temps == []) :
                    newdf_temp = df
                else : 
                    filtrage= (df["NUIT"].astype(str).isin(selected_temps))
                    newdf_temp = df[filtrage]
            
        else : 
            selected_temps.append(checkbox_var_temp.get())
            print(selected_temps)
            if(selected_users != []):
                filtrage= (df["NUIT"].astype(str).isin(selected_temps) & df["UserID"].isin(selected_users))
            else:
                filtrage= (df["NUIT"].astype(str).isin(selected_temps))
            newdf_temp = df[filtrage]

        fillTable(newdf_temp, "Données")

    
    workbook3 = Workbook()
    sheet = workbook3.active
    sheet.append(["Time", "Activity"])
    sorted_stats = sorted(stats, key=lambda x: x.activity[0][0])
    
    for instance in sorted_stats:
        sheetTitle = instance.id + "_" + instance.NUIT
        donnees = instance.activity[:]
        sheet.title = str(sheetTitle)
        for i, donnee in enumerate(donnees):
            heure = donnee[0]
            activite = donnee[1]
            sheet.cell(row=i + 1, column=1, value=heure)
            sheet.cell(row=i + 1, column=2, value=activite)
        sheet = workbook3.create_sheet()
    
    workbook3.save(filename="test_activite.xlsx")
    workbook3.close()

    combined_df = pd.DataFrame()
    excel_file = 'test_activite.xlsx'
    xls = pd.ExcelFile(excel_file)
    for sheet_name in xls.sheet_names:
        dfExcel = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)

        if not dfExcel.empty : 
            dfExcel.columns = ['Time', str(sheet_name)]

            # Combinaison des données dans le DataFrame global
            combined_df = pd.concat([combined_df, dfExcel], ignore_index=True)

    distinct_hours = combined_df['Time'].unique()
    final_data = pd.DataFrame({'Time': distinct_hours})

    for sheet_name in xls.sheet_names:
        data = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)

        if not data.empty : 
            data.columns = ['Time', str(sheet_name) ]

            final_data = final_data.merge(data, on='Time', how='left')
    

    # Remplacer les valeurs manquantes par 0
    final_data.fillna(0, inplace=True)

    new_file_path = 'CombinedActivity.xlsx'
    workbook4 = Workbook()
    worksheet = workbook4.active
    worksheet.title = 'Feuille_combinee'

    for row in dataframe_to_rows(final_data, index=False, header=True):
        worksheet.append(row)

    workbook4.save(new_file_path)

    workbook5 = Workbook()
    sheet = workbook5.active
    data = dataframe_to_rows(df_sorted, index = False, header=True)

    for row in data:
        sheet.append(row)
    workbook5.save("DécoupageNuit.xlsx")

    df = pd.DataFrame(pd.read_excel("test.xlsx"))

    buttonSelectionFichier.destroy()


    style = ttk.Style()
    
    style.theme_use("default")
    
    style.configure("Treeview",
                            background="#2a2d2e",
                            foreground="white",
                            rowheight=25,
                            fieldbackground="#343638",
                            bordercolor="#343638",
                            borderwidth=0)
    style.map('Treeview', background=[('selected', '#565b5e')])
    
    style.configure("Treeview.Heading",
                            background="#52a57c",
                            foreground="white",
                            relief="flat")
    style.map("Treeview.Heading", background=[('active', '#52a57c')])




    def fillTable(newdf, tab):
        
        df_list=list(newdf)
        df_rset=newdf.round(2).to_numpy().tolist()

        df_tree = ttk.Treeview(main_window.tabView.tab(tab), columns=df_list)
        for i in df_tree.get_children():
            df_tree.delete(i)
        df_tree["show"] = "headings"
        df_tree.grid(row=3, column=1, columnspan=3, padx=(20, 20), pady=(10, 10), sticky="ew")

        for i in df_list:
            df_tree.column(i,width=75,anchor='c')
            df_tree.heading(i,text=i)
        for dt in df_rset:
            v=[r for r in dt]
            df_tree.insert('','end', values=v)
    
    fillTable(df, "Données")

    optionmenu1_var = customtkinter.StringVar(value=tri[0])
    optionmenu2_var = customtkinter.StringVar(value=triSE[0])

    optionmenu_user = customtkinter.CTkOptionMenu(master=main_window.tabView.tab("Données"), dynamic_resizing=False, values = tri, command=optionmenu_triSFI_callback, variable=optionmenu1_var)
    optionmenu_user.grid(row=1, column=1, padx=(100, 100), pady=(20, 20), sticky="nsew")
    optionmenu_2 = customtkinter.CTkOptionMenu(master=main_window.tabView.tab("Données"), dynamic_resizing=False, values = triSE, command=optionmenu_triSE_callback, variable=optionmenu2_var)
    optionmenu_2.grid(row=1, column=2,  padx=(100, 100), pady=(20, 20), sticky="nsew")
    optionmenu_3 = customtkinter.CTkOptionMenu(master=main_window.tabView.tab("Données"), dynamic_resizing=False, values = users, command=optionmenu_triSFI_callback, variable=optionmenu1_var)
    optionmenu_3.grid(row=1, column=3,  padx=(100, 100), pady=(20, 20), sticky="nsew")
   
    zoom = 0.8


    img = Image.open("sfi_all_means.png")

    pixels_x, pixels_y = tuple([int(zoom * x)  for x in img.size])
    
    img = ImageTk.PhotoImage(img.resize((pixels_x, pixels_y)))
    panel = Label(main_window.tabView.tab("Stats"), image = img)
    panel.photo = img
    panel.grid(column=1, row=1, padx=(20, 20), pady=(10, 10), sticky="nsew")

    checkbox_var_users = customtkinter.StringVar(value=users[0])

    checkbox_frame = customtkinter.CTkScrollableFrame(main_window.tabView.tab("Données"))
    checkbox_frame.grid(row=8, column=1, padx=(20, 20), pady=(10, 10), sticky="nsew")
    checkbox_frame.grid_columnconfigure(0, weight=1)
    titre_checkbox = customtkinter.CTkLabel(master=checkbox_frame, text="Filtrer par user(s) : ", font=customtkinter.CTkFont(size=20))
    titre_checkbox.grid(row=1, column=1,  padx=(20, 20), pady=(10, 10), sticky="nsew")

    i = 1
    for user in users [1:] : 
        checkbox = customtkinter.CTkCheckBox(master=checkbox_frame, text=user, variable=checkbox_var_users, offvalue = user, onvalue = user, command=checkbox_user_callback)
        checkbox.grid(row=i +1, column=1, padx=(20, 20), pady=(10, 10), sticky="nsew")
        i = i+1
    

    def radiobuttonSelection():

        if(radio_var.get() == "Histogramme (moyennes)"):
            
            if(radio2_var.get() == "SFI"):
                img = Image.open("sfi_all_means.png")
            elif(radio2_var.get() == "Sleep efficiency (%)") : 
                img = Image.open("sleep_efficiency_all_means.png")
            elif(radio2_var.get() == "sleep_latency"): 
                img = Image.open("sleep_latency_all_means.png")
            elif(radio2_var.get() == "Tout"):
                img = Image.open("all_means.png")

            img = ImageTk.PhotoImage(img.resize((pixels_x, pixels_y)))
            panel = Label(main_window.tabView.tab("Stats"), image = img)
            panel.photo = img
            panel.grid(column=1, row=1, padx=(20, 20), pady=(10, 10), sticky="nsew")
        elif(radio_var.get() == "autre"): 
            img = Image.open("testPlot4.png")
            img = ImageTk.PhotoImage(img.resize((pixels_x, pixels_y)))
            panel = Label(main_window.tabView.tab("Stats"), image = img)
            panel.photo = img
            panel.grid(column=1, row=1, padx=(20, 20), pady=(10, 10), sticky="nsew")
        elif(radio_var.get() == "Boîtes à moustache"): 
            if(radio2_var.get() == "SFI"):
                img = Image.open("boxplotSFI.png")
            elif(radio2_var.get() == "Sleep efficiency (%)") : 
                    img = Image.open("boxplotSE.png")
            elif(radio2_var.get() == "sleep_latency"):
                    img = Image.open("boxplotSL.png")

            img = ImageTk.PhotoImage(img.resize((pixels_x, pixels_y)))
            panel = Label(main_window.tabView.tab("Stats"), image = img)
            panel.photo = img
            panel.grid(column=1, row=1, padx=(20, 20), pady=(10, 10), sticky="nsew")

    radiobutton_frame_tab_2 = customtkinter.CTkFrame(main_window.tabView.tab("Stats"))

    radiobutton_frame_tab_2.grid(row=1, column=4, padx=(20, 20), pady=(10, 10), sticky="nsew")
    radio_var = tk.StringVar(value = "Histogramme (moyennes)")

    radiobuttonHistogrammeTab2 = customtkinter.CTkRadioButton(master=radiobutton_frame_tab_2, variable=radio_var, value = "Histogramme (moyennes)", text="Histogramme (moyennes)", command=radiobuttonSelection)
    radiobuttonHistogrammeTab2.grid(row=2, column=0, pady=(20, 0), padx=20, sticky="nsew")

    radobutton3 = customtkinter.CTkRadioButton(master=radiobutton_frame_tab_2, variable=radio_var, value = "Boîtes à moustache", text = "Boîtes à moustache", command=radiobuttonSelection)
    radobutton3.grid(row=4, column=0, pady=20, padx=20, sticky="nsew")
    
    titre_radioframe = customtkinter.CTkLabel(master=radiobutton_frame_tab_2, text="Type de graphe : ", font=customtkinter.CTkFont(size=20))
    titre_radioframe.grid(row=1, column=0,  padx=(20, 20), pady=(10, 10), sticky="nsew")

    radiobutton_frame2 = customtkinter.CTkFrame(main_window.tabView.tab("Stats"))

    radiobutton_frame2.grid(row=3, column=4, padx=(20, 20), pady=(10, 10), sticky="nsew")
    radio2_var = tk.StringVar(value="SFI")
    radiobutton_sleep_efficiency = customtkinter.CTkRadioButton(master=radiobutton_frame2, variable=radio2_var, value = "Sleep efficiency (%)", text="Sleep efficiency (%)", command=radiobuttonSelection)
    radiobutton_sleep_efficiency.grid(row=2, column=0, pady=(20, 0), padx=20, sticky="nsew")
    radiobutton_SFI = customtkinter.CTkRadioButton(master=radiobutton_frame2, variable=radio2_var, value = "SFI", text="SFI", command=radiobuttonSelection)
    radiobutton_SFI.grid(row=3, column=0, pady=(20, 0), padx=20, sticky="nsew")
    radiobutton_sleep_latency = customtkinter.CTkRadioButton(master=radiobutton_frame2, variable=radio2_var, value = "sleep_latency", text="sleep_latency", command=radiobuttonSelection)
    radiobutton_sleep_latency.grid(row=4, column=0, pady=(20, 0), padx=20, sticky="nsew")
    radiobutton_tout = customtkinter.CTkRadioButton(master=radiobutton_frame2, variable=radio2_var, value = "Tout", text="Tout", command=radiobuttonSelection)
    radiobutton_tout.grid(row=2, column=1, pady=(20, 0), padx=20, sticky="nsew")


    titre_radioframe2 = customtkinter.CTkLabel(master=radiobutton_frame2, text="Variable : ", font=customtkinter.CTkFont(size=20))
    titre_radioframe2.grid(row=1, column=0,  padx=(20, 20), pady=(10, 10), sticky="nsew")
    

    checkbox_var_temp = customtkinter.StringVar(value=temp[0])

    checkbox2_frame = customtkinter.CTkScrollableFrame(main_window.tabView.tab("Données"))
    checkbox2_frame.grid(row=8, column=2, padx=(20, 20), pady=(10, 10), sticky="nsew")
    checkbox2_frame.grid_columnconfigure(1, weight=1)

    titre_checkbox2 = customtkinter.CTkLabel(master=checkbox2_frame, text="Filtrer par nuit : ", font=customtkinter.CTkFont(size=20))
    titre_checkbox2.grid(row=1, column=1,  padx=(20, 20), pady=(10, 10), sticky="nsew")

    i = 1
    for temperature in temp[1:] : 
        checkbox = customtkinter.CTkCheckBox(master=checkbox2_frame, text=temperature, variable=checkbox_var_temp, offvalue = temperature, onvalue = temperature, command=checkbox_temp_callback)
        checkbox.grid(row=i +1, column=1, padx=(20, 20), pady=(10, 10), sticky="nsew")
        i = i+1
    
    statsvar = customtkinter.StringVar(value=moy)
    fillTable(moydf, "Stats")



    zoom = 0.8


    img = Image.open("userHour.png")

    pixels_x, pixels_y = tuple([int(zoom * x)  for x in img.size])
    
    img = ImageTk.PhotoImage(img.resize((pixels_x, pixels_y)))
    panel = Label(main_window.tabView.tab("Découpage nuit"), image = img)
    panel.photo = img
    panel.grid(column=1, row=1, padx=(20, 20), pady=(10, 10), sticky="nsew")


    def radiobuttonSelection2():

        plotHourbyUser(checkbox3_var.get(), radio4_var.get())
        img = Image.open("userHour.png")

        img = ImageTk.PhotoImage(img.resize((pixels_x, pixels_y)))
        panel = Label(main_window.tabView.tab("Découpage nuit"), image = img)
        panel.photo = img
        panel.grid(column=1, row=1, padx=(20, 20), pady=(10, 10), sticky="nsew")

    radio3_frame = customtkinter.CTkScrollableFrame(main_window.tabView.tab("Découpage nuit"))
    radio3_frame.grid_columnconfigure(1, weight=1)

    radio3_frame.grid(row=1, column=4, padx=(20, 20), pady=(10, 10), sticky="nsew")
    checkbox3_var = tk.StringVar(value = "")

    
    def radiobuttonSelection3 () : 
        plotHourbyUser(checkbox3_var.get(), radio4_var.get())
        img = Image.open("userHour.png")

        img = ImageTk.PhotoImage(img.resize((pixels_x, pixels_y)))
        panel = Label(main_window.tabView.tab("Découpage nuit"), image = img)
        panel.photo = img
        panel.grid(column=1, row=1, padx=(20, 20), pady=(10, 10), sticky="nsew")
    


    for user in users [1:] : 
        radio = customtkinter.CTkRadioButton(master=radio3_frame, text=user, variable=checkbox3_var, value = user,command=radiobuttonSelection3)
        radio.grid(row=i +1, column=1, padx=(20, 20), pady=(10, 10), sticky="nsew")
        i = i+1


    radio4_var = tk.StringVar(value="Tout")

    checkbox_frame.grid_rowconfigure(2, weight=0)

    radiobutton_frame3 = customtkinter.CTkScrollableFrame(main_window.tabView.tab("Découpage nuit"))

    radiobutton_frame3.grid(row=3, column=4, padx=(20, 20), pady=(10, 10), sticky="nsew")


    i = 1
    for temperature in temp[1:] : 
        print (temperature)
        radiobutton_nui_decoupage = customtkinter.CTkRadioButton(master=radiobutton_frame3, text=temperature, value = temperature, variable=radio4_var, command=radiobuttonSelection2)
        radiobutton_nui_decoupage.grid(row=i +1, column=1, padx=(20, 20), pady=(10, 10), sticky="nsew")
        i = i+1
    

    titre_radioframe2 = customtkinter.CTkLabel(master=radiobutton_frame3, text="Nuit: ", font=customtkinter.CTkFont(size=20))
    titre_radioframe2.grid(row=0, column=0,  padx=(10, 20), pady=(10, 10), sticky="nsew")
    
    i = 1
    for temperature in temp[1:] : 
        checkbox = customtkinter.CTkCheckBox(master=checkbox2_frame, text=temperature, variable=checkbox_var_temp, offvalue = temperature, onvalue = temperature, command=checkbox_temp_callback)
        checkbox.grid(row=i +1, column=1, padx=(20, 20), pady=(10, 10), sticky="nsew")
        i = i+1
    
    fillTable(moydfHour, "Découpage nuit")

    def plotHourbyUser(user, temperature) : 
        if(temperature == "Tout") : 
            colonnes = df_sorted.filter(like = user).columns.to_list()
            colonnesTout = df_sorted.filter(like = "_" ).columns.to_list()
        
        else: 
            colonnes = df_sorted.filter(like = user).filter(like=temperature).columns.to_list()
            colonnesTout = df_sorted.filter(like=temperature).columns.to_list()
        df_sorted['Heure'] = pd.to_datetime(df_sorted['Heure'], format='%H:%M:%S')

        meanActivity = df_sorted[colonnes].mean(axis=1)
        df_sorted["mean"] = meanActivity
        meanActivityAll = df_sorted[colonnesTout].mean(axis=1)


        df_sorted["meanAll"] = meanActivityAll


        df_sorted.reset_index(drop=False)
        df_sorted["Heure"] = df_sorted["Heure"].dt.strftime("%H:%M:%S")

        figUser = df_sorted.plot(x = "Heure", y = "meanAll", rot = 1, ylim=(0,50), xlabel = "Heure", ylabel = "Activité", label= "Moy tous les sujets", color = "red")
    
        if(temperature == "Tout") : 
            figUser2 = df_sorted.plot.bar(x = "Heure", y = "meanAll", rot = 1, ylim=(0,50), xlabel = "Heure", ylabel = "Activité", ax = figUser).get_figure()
        else :    
            figUser2 = df_sorted.plot.bar(x = "Heure", y = "mean", rot = 1, ylim=(0,50), xlabel = "Heure", ylabel = "Activité", label= user + " " + temperature, ax = figUser).get_figure()

        plt.locator_params(axis='x', nbins=7)
        plt.savefig("userHour.png")

    
    

def UploadAction():
    file_path = filedialog.askopenfilename()
    if file_path is not None:
        print (file_path)
        loadWorkbook(file_path)
        textSelectionFichier.set("Filtrage par sujet")






textSelectionFichier = StringVar()
textSelectionFichier.set("Sélectionner le fichier à traiter")
main_window.labelSelectionFichier = customtkinter.CTkLabel(master = main_window.tabView.tab("Données"), textvariable = textSelectionFichier, font=customtkinter.CTkFont(size=40, weight="bold"))
main_window.labelSelectionFichier.place(relx=.5, rely=.3, anchor="center")
buttonSelectionFichier = customtkinter.CTkButton(master = main_window.tabView.tab("Données"), text='Sélectionner...', command=UploadAction, width=200, height=50, font=customtkinter.CTkFont(size=20))
buttonSelectionFichier.place(relx=.5, rely=.6, anchor="center")




main_window.mainloop()